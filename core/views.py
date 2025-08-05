from datetime import timedelta
from decimal import Decimal
from django.shortcuts import render
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages

def login_view(request):
    if request.user.is_authenticated:
        return redirect('client_list')  # Или другой ваш главный URL
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            next_url = request.GET.get('next', 'client_list')  # Получаем параметр next или используем home
            return redirect(next_url)
        else:
            messages.error(request, 'Неверное имя пользователя или пароль')
    
    return render(request, 'registration/login.html')

@login_required
def logout_view(request):
    logout(request)
    messages.success(request, 'Вы успешно вышли из системы')
    return redirect('client_list')  # Или другой URL после выхода

@login_required
def protected_view(request):
    # Пример защищенного представления
    return render(request, 'protected.html')
# Create your views here.
from django.shortcuts import get_object_or_404, render
from django.http import JsonResponse
from .models import CashOperation, CashRegister, Client, Debt, Membership, MembershipType, Payment, Visit
from django.utils import timezone
from django.shortcuts import render
from .models import Client
from django.contrib.auth.decorators import login_required
@login_required
def client_list(request):
    # Получаем параметры поиска из GET-запроса
    search_name = request.GET.get('search_name', '')
    search_card = request.GET.get('search_card', '')
    
    # Начинаем с базового QuerySet
    clients = Client.objects.filter(is_active=True)
    
    # Применяем фильтры поиска
    if search_name:
        clients = clients.filter(full_name__icontains=search_name)
    if search_card:
        clients = clients.filter(card_id__icontains=search_card)
    
    # Сортируем результаты
    clients = clients.order_by('full_name')
    
    # Передаем в шаблон
    context = {
        'clients': clients,
        'title': 'Список клиентов',
        'search_name': search_name,
        'search_card': search_card
    }
    return render(request, 'client_list.html', context)

from django.db import transaction

def get_main_cash_register():
    """Получаем или создаем основную кассу"""
    cash_register, created = CashRegister.objects.get_or_create(
        name="Основная касса",
        defaults={'balance': 0}
    )
    return cash_register
# @login_required
# def client_list(request):
#     # Получаем всех активных клиентов, сортируем по ФИО
#     clients = Client.objects.filter(is_active=True).order_by('full_name')
    
#     # Передаем в шаблон
#     context = {
#         'clients': clients,
#         'title': 'Список клиентов'
#     }
#     return render(request, 'client_list.html', context)
from django.shortcuts import render, redirect
from .forms import CardSaleForm, ClientForm, MembershipForm, WithdrawForm
from django.contrib import messages

@login_required
def add_client(request):
    if request.method == 'POST':
        form = ClientForm(request.POST, request.FILES)
        if form.is_valid():
            client = form.save()
            messages.success(request, f'Клиент {client.full_name} успешно добавлен!')
            return redirect('client_detail', client.id)
    else:
        form = ClientForm()
    
    context = {
        'form': form,
        'title': 'Добавить нового клиента'
    }
    return render(request, 'core/client_form.html', context)


def check_access(request):
    if request.method == 'POST':
        card_id = request.POST.get('card_id')
        
        try:
            client = Client.objects.get(card_id=card_id, is_active=True)
            active_memberships = Membership.objects.filter(
                client=client,
                is_active=True,
                end_date__gte=timezone.now().date(),
                remaining_visits__gt=0
            ).order_by('end_date')
            
            if active_memberships.exists():
                membership = active_memberships.first()
                membership.remaining_visits -= 1
                membership.save()
                
                Visit.objects.create(
                    membership=membership,
                    success=True
                )
                
                return JsonResponse({
                    'access': True,
                    'client': client.full_name,
                    'remaining': membership.remaining_visits
                })
            else:
                Visit.objects.create(
                    client=client,
                    success=False,
                    reason="Нет активного абонемента"
                )
                return JsonResponse({
                    'access': False,
                    'reason': "Нет активного абонемента"
                })
                
        except Client.DoesNotExist:
            return JsonResponse({
                'access': False,
                'reason': "Клиент не найден"
            })
    
    return render(request, 'core/check_access.html')

from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.db.models import Sum

@transaction.atomic
@login_required
def client_detail(request, pk):
    client = get_object_or_404(Client, pk=pk)
    cash_register = get_main_cash_register()
    total_debt = Debt.objects.filter(
        client=client, 
        is_paid=False
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    active_memberships = Membership.objects.filter(client=client, is_active=True)
    # Обработка платежа по долгу
    # if request.method == 'POST' and 'add_membership' in request.POST:
    #     form = MembershipForm(request.POST)
    #     if form.is_valid():
    #         membership = form.save(commit=False)
    #         membership.client = client
    #         membership.purchase_date = timezone.now().date()
    #         membership.end_date = membership.start_date + timedelta(days=membership.membership_type.validity_days)
    #         membership.original_price = membership.membership_type.price
            
    #         discount = form.cleaned_data.get('discount_amount', 0)
    #         paid = form.cleaned_data.get('paid_amount', 0)
    #         payment_method = form.cleaned_data.get('payment_method', 'cash')
            
    #         # Проверки
    #         if discount > membership.original_price:
    #             messages.error(request, 'Скидка не может быть больше стоимости абонемента!')
    #             return redirect('client_detail', pk=client.id)
            
    #         if paid < 0:
    #             messages.error(request, 'Оплаченная сумма не может быть отрицательной!')
    #             return redirect('client_detail', pk=client.id)
            
    #         # Сохраняем абонемент
    #         membership.discount_amount = discount
    #         membership.paid_amount = paid
    #         membership.created_by = request.user
    #         membership.save()
            
    #         # Создаем платеж
    #         payment = Payment.objects.create(
    #             membership=membership,
    #             amount=paid,
    #             payment_method=payment_method,
    #             discount_amount=discount,
    #             created_by=request.user
    #         )
            
    #         # Записываем операцию в кассу (если это не долг)
    #         if paid > 0:
    #             CashOperation.objects.create(
    #                 cash_register=cash_register,
    #                 operation_type='income',
    #                 amount=paid,
    #                 payment=payment,
    #                 notes=f"Оплата абонемента {membership.membership_type.name}",
    #                 created_by=request.user
    #             )
    #             messages.success(request, f'В кассу зачислено {paid} сом')
            
    #         # Обработка долга
    #         required_payment = membership.original_price - discount
    #         if paid < required_payment:
    #             debt_amount = required_payment - paid
    #             membership.debt_amount = debt_amount
    #             membership.save()
                
    #             Debt.objects.create(
    #                 client=client,
    #                 amount=debt_amount,
    #                 membership=membership,
    #                 created_by=request.user,
    #                 description=f"Долг за абонемент {membership.membership_type.name}"
    #             )
    #             messages.warning(request, f'Сформирован долг: {debt_amount} сом')
            
    #         messages.success(request, 'Абонемент успешно добавлен!')
    #         return redirect('client_detail', pk=client.id)
    
    if request.method == 'POST' and 'pay_debt' in request.POST:
        amount = Decimal(request.POST.get('debt_payment_amount', 0))
        debt_id = request.POST.get('debt_id')
        payment_method = request.POST.get('payment_method', 'cash')
        
        if amount <= 0:
            messages.error(request, 'Сумма платежа должна быть положительной!')
            return redirect('client_detail', pk=client.id)
        
        try:
            debt = Debt.objects.get(id=debt_id, client=client, is_paid=False)
            
            if amount > debt.amount:
                messages.warning(request, f'Сумма платежа ({amount} сом) превышает сумму долга ({debt.amount} сом). Будет зачислено {debt.amount} сом.')
                amount = debt.amount
            
            # Создаем запись о платеже
            payment = Payment.objects.create(
                membership=debt.membership,
                amount=amount,
                payment_method=payment_method,
                is_debt_payment=True,
                notes=f"Погашение долга #{debt.id}",
                created_by=request.user
            )
            
            # Записываем операцию в кассу
            CashOperation.objects.create(
                cash_register=cash_register,
                operation_type='income',
                amount=amount,
                payment=payment,
                notes=f"Погашение долга #{debt.id}",
                created_by=request.user
            )
            
            # Обновляем долг
            debt.amount -= amount
            if debt.amount <= 0:
                debt.is_paid = True
                messages.success(request, 'Долг полностью погашен!')
            else:
                messages.success(request, f'Частичное погашение долга. Остаток: {debt.amount} сом')
            
            debt.save()
            messages.success(request, f'В кассу зачислено {amount} сом')
            
        except Debt.DoesNotExist:
            messages.error(request, 'Долг не найден или уже погашен!')
        
        return redirect('client_detail', pk=client.id)
    # if request.method == 'POST' and 'pay_debt' in request.POST:
    #     amount = Decimal(request.POST.get('debt_payment_amount', 0))
    #     debt_id = request.POST.get('debt_id')  # Добавляем выбор конкретного долга
    
    #     if amount <= 0:
    #         messages.error(request, 'Сумма платежа должна быть положительной!')
    #         return redirect('client_detail', pk=client.id)
    
    #     try:
    #         debt = Debt.objects.get(id=debt_id, client=client, is_paid=False)
        
    #         if amount > debt.amount:
    #             messages.warning(request, f'Сумма платежа ({amount} сом) превышает сумму долга ({debt.amount} сом). Будет зачислено {debt.amount} сом.')
    #             amount = debt.amount
        
    #     # Создаем запись о платеже
    #         Payment.objects.create(
    #         membership=debt.membership if debt.membership else None,
    #         amount=amount,
    #         payment_method='cash',  # или получать из формы
    #         payment_date=timezone.now(),
    #         is_debt_payment=True,
    #         notes=f"Погашение долга #{debt.id}",
    #         created_by=request.user
    #     )
        
    #     # Обновляем сумму долга
    #         debt.amount -= amount
    #         if debt.amount <= 0:
    #             debt.is_paid = True
    #             messages.success(request, 'Долг полностью погашен!')
    #         else:
    #             messages.success(request, f'Частичное погашение долга. Остаток: {debt.amount} сом')
        
    #         debt.save()
        
    #     except Debt.DoesNotExist:
    #         messages.error(request, 'Долг не найден или уже погашен!')
    
    #     return redirect('client_detail', pk=client.id)
    # if request.method == 'POST' and 'pay_debt' in request.POST:
    #     amount = Decimal(request.POST.get('debt_payment_amount', 0))
    #     if amount > 0:
    #         # Логика погашения долга
    #         Payment.objects.create(
    #             # client=client,
    #             amount=amount,
    #             payment_method='debt_payment',
    #             notes="Погашение долга"
    #         )
    #         messages.success(request, f'Платеж на {amount} сом зачислен в счет долга')
    #         return redirect('client_detail', pk=client.id)
    
    # Активные абонементы
    active_memberships = Membership.objects.filter(
        client=client,
        is_active=True,
        end_date__gte=timezone.now().date()
    ).order_by('-start_date')
    
    # Все абонементы (для истории)
    all_memberships = Membership.objects.filter(client=client).order_by('-purchase_date')
    
    # Общее количество посещений
    total_visits = Visit.objects.filter(
        membership__client=client,
        success=True
    ).count()
    
    # Обработка отметки посещения
    if request.method == 'POST' and 'mark_visit' in request.POST:
        membership_id = request.POST.get('membership_id')
        try:
            membership = Membership.objects.get(id=membership_id, client=client)
            if membership.remaining_visits > 0:
                membership.remaining_visits -= 1
                membership.save()
                Visit.objects.create(
                    membership=membership,
                    success=True
                )
                messages.success(request, 'Посещение успешно отмечено!')
            else:
                Visit.objects.create(
                    membership=membership,
                    success=False,
                    reason="Не осталось посещений"
                )
                messages.warning(request, 'Не осталось посещений по этому абонементу!')
        except Membership.DoesNotExist:
            messages.error(request, 'Абонемент не найден!')
        return redirect('client_detail', pk=client.id)
    
    if request.method == 'POST' and 'add_membership' in request.POST:
        form = MembershipForm(request.POST)
        if form.is_valid():
            membership = form.save(commit=False)
            membership.client = client
            membership.purchase_date = timezone.now().date()
            membership.end_date = membership.start_date + timedelta(days=membership.membership_type.validity_days)
            membership.created_by = request.user
        # Устанавливаем original_price из цены типа абонемента
            membership.original_price = membership.membership_type.price
        
        # Получаем данные из формы
            discount = form.cleaned_data.get('discount_amount', 0)
            paid = form.cleaned_data.get('paid_amount', 0)
        
        # Рассчитываем сколько должно быть оплачено с учетом скидки
            required_payment = membership.original_price - discount
        
        # Проверки
            if discount > membership.original_price:
                messages.error(request, 'Скидка не может быть больше стоимости абонемента!')
                return redirect('client_detail', pk=client.id)
        
            if paid < 0:
                messages.error(request, 'Оплаченная сумма не может быть отрицательной!')
                return redirect('client_detail', pk=client.id)
        
        # Сохраняем основные данные абонемента
            membership.discount_amount = discount
            membership.paid_amount = paid
            
        
        # Определяем есть ли долг
            has_debt = paid < required_payment
            if has_debt:
                membership.debt_amount = required_payment - paid
        
        # Сначала сохраняем абонемент
            membership.save()
        
        # Теперь создаем долг (если есть), после сохранения абонемента
            if has_debt:
                Debt.objects.create(
                client=client,
                amount=membership.debt_amount,
                membership=membership,  # Теперь membership имеет id
                created_by=request.user,
                description=f"Долг за абонемент {membership.membership_type.name}"
            )
                messages.warning(request, f'Сформирован долг: {membership.debt_amount} сом')
        
        # Создаем запись о платеже
            payment =  Payment.objects.create(
            membership=membership,
            amount=paid,
            payment_method='cash',
            discount_amount=discount,
            created_by=request.user
        )
        # Записываем операцию в кассу
            CashOperation.objects.create(
                cash_register=cash_register,
                operation_type='income',
                amount=paid,
                payment=payment,
                notes=f"Купили абонимент$",
                created_by=request.user
            )
            
            messages.success(request, 'Абонемент успешно добавлен!')
            return redirect('client_detail', pk=client.id)
    # if request.method == 'POST' and 'add_membership' in request.POST:
    #     form = MembershipForm(request.POST)
    #     if form.is_valid():
    #         membership = form.save(commit=False)
    #         membership.client = client
    #         membership.purchase_date = timezone.now().date()
    #         membership.end_date = membership.start_date + timedelta(days=membership.membership_type.validity_days)
            
    #         # Устанавливаем original_price из цены типа абонемента
    #         membership.original_price = membership.membership_type.price
            
    #         # Рассчитываем discount_amount и paid_amount
    #         discount = form.cleaned_data.get('discount_amount', 0)
    #         paid = form.cleaned_data.get('paid_amount', membership.original_price - discount)
            
    #         membership.discount_amount = discount
    #         membership.paid_amount = paid
    #         membership.created_by = request.user
    #         # Если оплачено меньше, чем цена со скидкой - создаем долг
    #         if paid < (membership.original_price - discount):
    #             membership.debt_amount = (membership.original_price - discount) - paid
            
    #         membership.save()
            
    #         # Создаем запись о платеже
    #         Payment.objects.create(
    #             membership=membership,
    #             amount=paid,
    #             payment_method='cash',
    #             discount_amount=discount,
    #             created_by=request.user  # Передаем текущего пользователя
    #         )
            
    #         messages.success(request, 'Абонемент успешно добавлен!')
    #         return redirect('client_detail', pk=client.id)
    else:
        form = MembershipForm(initial={
            'paid_amount': 0  # Или другая логика начального значения
        })
    # if request.method == 'POST' and 'add_membership' in request.POST:
    #     form = MembershipForm(request.POST)
    #     if form.is_valid():
    #         membership = form.save(commit=False)
    #         membership.client = client
    #         membership.purchase_date = timezone.now().date()
    #         membership.end_date = membership.start_date + timedelta(days=membership.membership_type.validity_days)
            
    #         original_price = membership.membership_type.price
    #         discount = form.cleaned_data.get('discount_amount', 0)
    #         paid = form.cleaned_data['paid_amount']
            
    #         # Проверяем, чтобы скидка не была больше цены
    #         if discount > original_price:
    #             messages.error(request, 'Скидка не может быть больше стоимости абонемента!')
    #             return redirect('client_detail', pk=client.id)
            
    #         membership.discount_amount = discount
    #         membership.paid_amount = paid
            
    #         # Сохраняем абонемент (в save() автоматически рассчитается долг)
    #         membership.save()
            
    #         # Создаем запись о платеже
    #         Payment.objects.create(
    #             membership=membership,
    #             amount=paid,
    #             payment_method='cash',
    #             discount_amount=discount
    #         )
            
    #         messages.success(request, 'Абонемент успешно добавлен!')
    #         return redirect('client_detail', pk=client.id)
    # else:
    #     form = MembershipForm()
    # if request.method == 'POST' and 'add_membership' in request.POST:
    #     form = MembershipForm(request.POST)
    #     if form.is_valid():
    #         membership = form.save(commit=False)
    #         membership.client = client
    #         membership.purchase_date = timezone.now().date()
    #         membership.end_date = membership.start_date + timedelta(days=membership.membership_type.validity_days)
    #         membership.save()
            
    #         # Создаем платеж
    #         Payment.objects.create(
    #             membership=membership,
    #             amount=membership.membership_type.price,
    #             payment_method='cash'  # или из формы
    #         )
            
    #         messages.success(request, 'Абонемент успешно добавлен!')
    #         return redirect('client_detail', pk=client.id)
    # else:
    #     form = MembershipForm()
    active_membership = Membership.objects.filter(
        client=client, 
        is_active=True
    ).first() 
    context = {
        'client': client,
        'total_debt': total_debt,
         'active_membership': active_membership,
        'active_memberships': active_memberships,
        'all_memberships': all_memberships,
        'total_visits': total_visits,
         'form': form,
    }
    return render(request, 'core/client_detail.html', context)

from django.views.generic import ListView
@login_required
@transaction.atomic
def sell_card(request):
    cash_register = get_main_cash_register()
    
    if request.method == 'POST' and 'sell_card' in request.POST:
        print("__________________________")
        form = CardSaleForm(request.POST)
        if form.is_valid():
            card_sale = form.save(commit=False)
            card_sale.sold_by = request.user
            card_sale.save()
             # Создаем запись о платеже
            payment =  Payment.objects.create(
           
            amount=card_sale.price,
            payment_method='cash',
            notes = 'продажа карты',
            created_by=request.user )
            # Записываем операцию в кассу
            CashOperation.objects.create(
                cash_register=cash_register,
                operation_type='income',
                amount=card_sale.price,
                payment=payment,
                notes=f"Продажа карты #{card_sale.card_number}",
                created_by=request.user
            )
            
            messages.success(request, f'Карта #{card_sale.card_number} продана за {card_sale.price} сом')
            return redirect('sell_card')
    else:
        form = CardSaleForm()
    
    withdraw_form = WithdrawForm()
    last_operations = CashOperation.objects.order_by('-created_at')[:10]
    
    context = {
        'form': form,
        'withdraw_form': withdraw_form,
        'cash_register': cash_register,
        'last_operations': last_operations,
    }
    return render(request, 'core/sell_card.html', context)
# @login_required
# @transaction.atomic
# def sell_card(request):
#     if request.method == 'POST':
#         form = CardSaleForm(request.POST)
#         if form.is_valid():
#             card_sale = form.save(commit=False)
#             card_sale.sold_by = request.user
#             card_sale.save()
            
#             # Записываем операцию в кассу
#             cash_register = get_main_cash_register()
#             CashOperation.objects.create(
#                 cash_register=cash_register,
#                 operation_type='income',
#                 amount=card_sale.price,
#                 card_sale=card_sale,
#                 notes=f"Продажа карты #{card_sale.card_number}",
#                 created_by=request.user
#             )
            
#             messages.success(request, f'Карта #{card_sale.card_number} продана за {card_sale.price} сом')
#             return redirect('sell_card')
#     else:
#         form = CardSaleForm()
    
#     return render(request, 'core/sell_card.html', {'form': form})

class CashOperationsListView(ListView):
    model = CashOperation
    template_name = 'core/cash_operations.html'
    context_object_name = 'operations'
    paginate_by = 50
    
    def get_queryset(self):
        return CashOperation.objects.select_related(
            'cash_register', 'payment', 'created_by'
        ).order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['cash_register'] = get_main_cash_register()
        return context

from django.views.generic import TemplateView
from django.utils import timezone
from datetime import timedelta
from openpyxl import Workbook
from django.http import HttpResponse

class ReportsView(TemplateView):
    template_name = 'core/reports.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Устанавливаем даты по умолчанию (последние 30 дней)
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=30)
        
        context['start_date'] = start_date.strftime('%Y-%m-%d')
        context['end_date'] = end_date.strftime('%Y-%m-%d')
        return context
from django.http import JsonResponse

def membership_type_info(request, pk):
    membership_type = get_object_or_404(MembershipType, pk=pk)
    data = {
        'name': membership_type.name,
        'visits_number': membership_type.visits_number,
        'validity_days': membership_type.validity_days,
        'price': str(membership_type.price),
    }
    return JsonResponse(data)


from django.http import HttpResponse
from openpyxl import Workbook
from datetime import timedelta
from django.utils import timezone
from datetime import timedelta

@login_required
def freeze_membership(request, pk):
    membership = get_object_or_404(Membership, pk=pk)
    if request.method == 'POST':
        # Нельзя заморозить уже завершенный абонемент
        if membership.end_date < timezone.now().date():
            messages.error(request, 'Нельзя заморозить просроченный абонемент')
            return redirect('client_detail', pk=membership.client.id)
            
        membership.is_frozen = True
        membership.freeze_start = timezone.now().date()
        membership.save()
        
        messages.success(request, 'Абонемент заморожен. Срок действия будет продлен после разморозки.')
        return redirect('client_detail', pk=membership.client.id)


@login_required
def unfreeze_membership(request, pk):
    membership = get_object_or_404(Membership, pk=pk)
    if request.method == 'POST':
        if not membership.is_frozen:
            messages.warning(request, 'Абонемент не заморожен')
            return redirect('client_detail', pk=membership.client.id)
            
        # Рассчитываем дни заморозки
        freeze_days = (timezone.now().date() - membership.freeze_start).days
        membership.freeze_days = freeze_days
        membership.is_frozen = False
        membership.save()  # В save() автоматически продлится срок
        
        messages.success(request, f'Абонемент разморожен. Срок продлен на {freeze_days} дней.')
        return redirect('client_detail', pk=membership.client.id)
# def freeze_membership(request, pk):
#     membership = get_object_or_404(Membership, pk=pk)
#     if request.method == 'POST':
#         membership.is_frozen = True
#         membership.save()
#         messages.success(request, 'Абонемент заморожен')
#         return redirect('client_detail', pk=membership.client.id)

# def unfreeze_membership(request, pk):
#     membership = get_object_or_404(Membership, pk=pk)
#     if request.method == 'POST':
#         membership.is_frozen = False
#         membership.end_date += timedelta(days=membership.frozen_days)
#         membership.save()
#         messages.success(request, 'Абонемент разморожен')
#         return redirect('client_detail', pk=membership.client.id)
@login_required
def refund_membership(request, pk):
    membership = get_object_or_404(Membership, pk=pk)
    PRICE_PER_VISIT = 400  # Цена за одно посещение при возврате
    
    if request.method == 'POST':
        # Проверяем, что абонемент еще активен
        if not membership.is_active:
            messages.error(request, 'Этот абонемент уже не активен')
            return redirect('client_detail', pk=membership.client.id)

        # Получаем все посещения по этому абонементу
        used_visits = Visit.objects.filter(membership=membership, success=True).count()
        total_visits = membership.membership_type.visits_number
        
        # Рассчитываем сумму, которую клиент "потратил" на посещения
        spent_amount = used_visits * PRICE_PER_VISIT
        
        # Рассчитываем сумму к возврату (не может быть отрицательной)
        refund_amount = max(membership.paid_amount - spent_amount, 0)
        
        # Если посещений больше чем в абонементе - ничего не возвращаем
        if used_visits >= total_visits:
            refund_amount = 0
        
        # Проверяем, что есть что возвращать
        if refund_amount <= 0:
            messages.warning(request, 
                f'Нет средств для возврата (использовано посещений: {used_visits}, '
                f'стоимость посещений: {used_visits * PRICE_PER_VISIT} сом)'
            )
            return redirect('client_detail', pk=membership.client.id)
        
        # Создаем запись о возврате
        payment = Payment.objects.create(
            membership=membership,
            amount=-refund_amount,
            payment_method='refund',
            notes=f"Возврат за абонемент (использовано {used_visits} посещений по 400 сом)",
            created_by=request.user
        )
        
        # Деактивируем абонемент
        membership.is_active = False
        membership.save()
        
        # Записываем операцию в кассу
        cash_register = get_main_cash_register()
        CashOperation.objects.create(
            cash_register=cash_register,
            operation_type='outcome',
            amount=refund_amount,
            payment=payment,
            notes=f"Возврат за абонемент {membership.membership_type.name}",
            created_by=request.user
        )
        
        # Формируем подробное сообщение
        message = (
            f"Возврат оформлен на {refund_amount} сом. "
            f"Использовано посещений: {used_visits} (по 400 сом = {used_visits * PRICE_PER_VISIT} сом). "
            f"Оплачено за абонемент: {membership.paid_amount} сом."
        )
        messages.success(request, message)
        
        return redirect('client_detail', pk=membership.client.id)
# @login_required
# def refund_membership(request, pk):
#     membership = get_object_or_404(Membership, pk=pk)
#     if request.method == 'POST':
#         # Проверяем, что абонемент еще активен
#         if not membership.is_active:
#             messages.error(request, 'Этот абонемент уже не активен')
#             return redirect('client_detail', pk=membership.client.id)

#         # Получаем все посещения по этому абонементу
#         used_visits = Visit.objects.filter(membership=membership, success=True).count()
#         total_visits = membership.membership_type.visits_number
        
#         # Рассчитываем стоимость одного посещения
#         price_per_visit = membership.paid_amount / total_visits
        
#         # Рассчитываем сумму к возврату
#         remaining_visits = total_visits - used_visits
#         refund_amount = round(price_per_visit * remaining_visits, 2)
        
#         # Проверяем, что есть что возвращать
#         if refund_amount <= 0:
#             messages.warning(request, 'Нет средств для возврата (все посещения использованы)')
#             return redirect('client_detail', pk=membership.client.id)
        
#         # Создаем запись о возврате
#         payment = Payment.objects.create(
#             membership=membership,
#             amount=-refund_amount,
#             payment_method='refund',
#             notes=f"Возврат за {remaining_visits} неиспользованных посещений",
#             created_by=request.user
#         )
        
#         # Деактивируем абонемент
#         membership.is_active = False
#         membership.save()
#         # Записываем операцию в кассу
#         cash_register = get_main_cash_register()
#         CashOperation.objects.create(
#                 cash_register=cash_register,
#                 operation_type='outcome',
#                 amount=refund_amount,
#                 payment=payment,
#                 notes=f"Возврат денег",
#                 created_by=request.user
#             )
#         # Формируем подробное сообщение
#         message = (
#             f"Возврат оформлен на {refund_amount} сом. "
#             f"Использовано посещений: {used_visits}/{total_visits}. "
#             f"Возвращено за {remaining_visits} посещений."
#         )
#         messages.success(request, message)
        
#         return redirect('client_detail', pk=membership.client.id)
# def refund_membership(request, pk):
#     membership = get_object_or_404(Membership, pk=pk)
#     if request.method == 'POST':
#         used_days = (timezone.now().date() - membership.start_date).days
#         total_days = (membership.end_date - membership.start_date).days
#         refund_amount = (membership.paid_amount / total_days) * (total_days - used_days)
        
#         # Создаем запись о возврате
#         Payment.objects.create(
#             membership=membership,
#             amount=-refund_amount,
#             payment_method='Возврат денег'
#         )
        
#         membership.is_active = False
#         membership.save()
#         messages.success(request, f'Возврат {refund_amount} сом оформлен')
def export_reports(request):
    # Получаем параметры фильтрации из GET-запроса
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Конвертируем строки в даты
    from django.utils.dateparse import parse_date
    start_date = parse_date(start_date)
    end_date = parse_date(end_date)
    
    # Фильтруем данные
    memberships = Membership.objects.filter(
        purchase_date__gte=start_date,
        purchase_date__lte=end_date
    ).select_related('client', 'membership_type')
    
    payments = Payment.objects.filter(
        payment_date__date__gte=start_date,
        payment_date__date__lte=end_date
    ).select_related('membership')
    
    # Создаем Excel файл
    wb = Workbook()
    
    # --- Лист с продажами абонементов ---
    ws_sales = wb.active
    ws_sales.title = "Продажи абонементов"
    headers_sales = [
        'Дата', 'Клиент', 'Тип абонемента', 
        'Полная цена', 'Скидка', 'Оплачено', 
        'Долг', 'Ответственный'
    ]
    ws_sales.append(headers_sales)
    
    # Переменные для сумм продаж
    total_full_price = 0
    total_discount = 0
    total_paid = 0
    total_debt = 0
    
    # Обработка обычных абонементов
    for m in memberships:
        # Преобразуем None в 0
        original_price = m.original_price or 0
        discount_amount = m.discount_amount or 0
        paid_amount = m.paid_amount or 0
        debt_amount = m.debt_amount or 0
        
        # Суммируем
        total_full_price += original_price
        total_discount += discount_amount
        total_paid += paid_amount
        total_debt += debt_amount
        
        ws_sales.append([
            m.purchase_date.strftime('%d.%m.%Y'),
            m.client.full_name,
            m.membership_type.name,
            original_price,
            discount_amount,
            paid_amount,
            debt_amount,
            m.created_by.username if m.created_by else 'Система'
        ])
    
    # Обработка одноразовых абонементов
    payment_aboniment = Payment.objects.filter(notes__icontains="Одноразовый абонемент").filter(
        payment_date__date__gte=start_date,
        payment_date__date__lte=end_date
    )

    for m in payment_aboniment:
        # Преобразуем None в 0
        amount = m.amount or 0
        discount_amount = m.discount_amount or 0
        
        # Суммируем (для одноразовых)
        total_full_price += amount
        total_discount += discount_amount
        total_paid += amount
        # Долг для одноразовых = 0
        
        ws_sales.append([
            m.payment_date.strftime('%d.%m.%Y %H:%M'),
            m.notes,
            m.get_payment_method_display(),
            amount,
            discount_amount,
            amount,
            0,
            m.created_by.username if m.created_by else 'Система'
        ])
    
    # Добавляем итоговую строку для продаж
    ws_sales.append([
        'ИТОГО:', '', '',
        total_full_price,
        total_discount,
        total_paid,
        total_debt,
        ''
    ])
    
    # --- Лист с платежами ---
    ws_payments = wb.create_sheet("Платежи")
    headers_payments = [
        'Дата', 'Клиент', 'Сумма', 
        'Способ оплаты', 'Тип платежа', 'Абонемент', "Кто внес платеж"
    ]
    ws_payments.append(headers_payments)
    
    # Переменные для сумм платежей
    total_payments = 0
    
    # Обработка платежей
    for p in payments:
        amount = p.amount or 0
        total_payments += amount
        
        ws_payments.append([
            p.payment_date.strftime('%d.%m.%Y %H:%M'),
            p.membership.client.full_name if p.membership else "долг",
            amount,
            p.get_payment_method_display(),
            'Погашение долга' if p.is_debt_payment else 'Оплата абонемента',
            f"{p.membership.membership_type.name} (до {p.membership.end_date.strftime('%d.%m.%Y')})" if p.membership else "Долг",
            p.created_by.username if p.created_by else 'Система'
        ])
    
    # Добавляем итоговую строку для платежей
    ws_payments.append([
        'ИТОГО:', '', 
        total_payments,
        '', '', '', ''
    ])
    
    # --- Стилизация итоговых строк ---
    from openpyxl.styles import Font, PatternFill
    
    # Стиль для итоговых строк
    bold_font = Font(bold=True)
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Стилизация листа продаж
    last_row_sales = ws_sales.max_row
    for col in range(1, 9):  # Колонки от A до H
        cell = ws_sales.cell(row=last_row_sales, column=col)
        cell.font = bold_font
        if col >= 4:  # Только для числовых колонок
            cell.fill = fill
    
    # Стилизация листа платежей
    last_row_payments = ws_payments.max_row
    for col in range(1, 8):  # Колонки от A до G
        cell = ws_payments.cell(row=last_row_payments, column=col)
        cell.font = bold_font
        if col == 3:  # Колонка с суммой
            cell.fill = fill
    
    # --- Настройка ширины столбцов ---
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        if col in ['D', 'E', 'F', 'G']:  # Числовые столбцы
            ws_sales.column_dimensions[col].width = 15
        else:
            ws_sales.column_dimensions[col].width = 25
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        if col == 'C':  # Столбец с суммой
            ws_payments.column_dimensions[col].width = 15
        else:
            ws_payments.column_dimensions[col].width = 25
    
    # Формируем HTTP-ответ
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=report_{start_date}_{end_date}.xlsx'
    wb.save(response)
    return response
# def export_reports(request):
#     # Получаем параметры фильтрации из GET-запроса
#     start_date = request.GET.get('start_date')
#     end_date = request.GET.get('end_date')
    
#     # Конвертируем строки в даты
#     from django.utils.dateparse import parse_date
#     start_date = parse_date(start_date)
#     end_date = parse_date(end_date)
    
#     # Фильтруем данные
#     memberships = Membership.objects.filter(
#         purchase_date__gte=start_date,
#         purchase_date__lte=end_date
#     ).select_related('client', 'membership_type')
    
#     payments = Payment.objects.filter(
#         payment_date__date__gte=start_date,
#         payment_date__date__lte=end_date
#     ).select_related('membership')
    
#     # Создаем Excel файл
#     wb = Workbook()
    
#     # Лист с продажами абонементов
#     ws_sales = wb.active
#     ws_sales.title = "Продажи абонементов"
#     ws_sales.append([
#         'Дата', 'Клиент', 'Тип абонемента', 
#         'Полная цена', 'Скидка', 'Оплачено', 
#         'Долг', 'Ответственный'
#     ])
    
#     for m in memberships:
#         ws_sales.append([
#             m.purchase_date.strftime('%d.%m.%Y'),
#             m.client.full_name,
#             m.membership_type.name,
#             m.original_price,
#             m.discount_amount,
#             m.paid_amount,
#             m.debt_amount,
#              m.created_by.username if m.created_by else 'Система'
#             # m.client.created_by.get_full_name() if m.client.created_by else ''
#         ])
#     payment_aboniment = Payment.objects.filter(notes__icontains = "Одноразовый абонемент").filter(
#         payment_date__date__gte=start_date,
#         payment_date__date__lte=end_date
#     )

#     for m in payment_aboniment:
#         ws_sales.append([
#             m.payment_date.strftime('%d.%m.%Y %H:%M'),
#             m.notes,
#             m.get_payment_method_display(),
#             m.amount,
#             m.discount_amount,
#             m.amount,
#             0,
#              m.created_by.username if m.created_by else 'Система'
#             # m.client.created_by.get_full_name() if m.client.created_by else ''
#         ])
#     # Лист с платежами
#     ws_payments = wb.create_sheet("Платежи")
#     ws_payments.append([
#         'Дата', 'Клиент', 'Сумма', 
#         'Способ оплаты', 'Тип платежа', 'Абонемент',"Кто внес платеж"
#     ])
    
#     for p in payments:
#         ws_payments.append([
#             p.payment_date.strftime('%d.%m.%Y %H:%M'),
#             p.membership.client.full_name if p.membership else "долг",
            
#             p.amount,
#             p.get_payment_method_display(),
#             'Погашение долга' if p.is_debt_payment else 'Оплата абонемента',
#             f"{p.membership.membership_type.name} (до {p.membership.end_date.strftime('%d.%m.%Y')})" if p.membership else "Долг" ,
#              p.created_by.username if p.created_by else 'Система'  # Новое поле p.created_by.get_full_name() if p.created_by else 'Система'  # Новое поле
#         ])
    
#     # Формируем HTTP-ответ
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = f'attachment; filename=report_{start_date}_{end_date}.xlsx'
#     wb.save(response)
#     return response        
# def export_reports(request):
#     period = request.GET.get('period', 'day')
#     date_from = request.GET.get('date_from')
#     date_to = request.GET.get('date_to')
#     # Определяем период
#     today = timezone.now().date()
#     if period == 'day':
#         date_from = today
#         date_to = today
#     elif period == 'week':
#         date_from = today - timedelta(days=today.weekday())
#         date_to = date_from + timedelta(days=6)
#     # ... аналогично для других периодов
    
#     # Собираем данные
#     memberships = Membership.objects.filter(
#         purchase_date__gte=date_from,
#         purchase_date__lte=date_to
#     )
    
#     # Создаем Excel файл
#     wb = Workbook()
#     ws = wb.active
    
#     # Заголовки
#     ws.append([
#         'Клиент', 'Тип абонемента', 'Дата покупки', 
#         'Цена', 'Скидка', 'Оплачено', 'Кто оформил'
#     ])
    
#     # Данные
#     for m in memberships:
#         ws.append([
#             m.client.full_name,
#             m.membership_type.name,
#             m.purchase_date.strftime('%d.%m.%Y'),
#             m.original_price,
#             m.discount_amount,
#             m.paid_amount,
#             m.client.created_by.username if m.client.created_by else ''
#         ])
    
#     # ... аналогично для других данных (продажи карт, посещения и т.д.)
    
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = f'attachment; filename=report_{period}.xlsx'
#     wb.save(response)
#     return response

from django.http import HttpResponse
from openpyxl import Workbook

def export_sales_report(request):
    # Получаем параметры фильтрации
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    # Фильтруем данные по продажам абонементов
    memberships = Membership.objects.filter(
        purchase_date__gte=date_from,
        purchase_date__lte=date_to
    ).select_related('client', 'membership_type')
    
    # Фильтруем данные по платежам
    payments = Payment.objects.filter(
        payment_date__gte=date_from,
        payment_date__lte=date_to
    ).select_related('membership__client')
    
    # Создаем Excel файл
    wb = Workbook()
    
    # --- Лист с продажами абонементов ---
    ws_sales = wb.active
    ws_sales.title = "Продажи абонементов"
    
    # Заголовки
    headers_sales = [
        'Дата', 'Клиент', 'Тип абонемента', 
        'Полная стоимость', 'Скидка', 'Оплачено', 
        'Долг', 'Ответственный'
    ]
    ws_sales.append(headers_sales)
    
    # Инициализируем переменные для сумм
    total_full_price = 0
    total_discount = 0
    total_paid_sales = 0
    total_debt = 0
    
    # Данные по продажам
    for m in memberships:
        full_price = m.membership_type.price or 0
        discount = m.discount_amount or 0
        paid = m.paid_amount or 0
        debt = m.debt_amount or 0
        
        total_full_price += full_price
        total_discount += discount
        total_paid_sales += paid
        total_debt += debt
        
        ws_sales.append([
            m.purchase_date.strftime('%d.%m.%Y'),
            m.client.full_name,
            m.membership_type.name,
            full_price,
            discount,
            paid,
            debt,
            request.user.get_full_name()
        ])
    
    # Добавляем итоги для продаж
    ws_sales.append([])  # Пустая строка
    ws_sales.append([
        'ИТОГО ПРОДАЖИ:', '', '',
        total_full_price,
        total_discount,
        total_paid_sales,
        total_debt,
        ''
    ])
    
    # --- Лист с платежами ---
    ws_payments = wb.create_sheet(title="Платежи")
    
    # Заголовки платежей
    headers_payments = [
        'Дата платежа', 'Клиент', 'Абонемент', 
        'Сумма платежа', 'Ответственный'
    ]
    ws_payments.append(headers_payments)
    
    # Инициализируем переменные для сумм платежей
    total_payments = 0
    
    # Данные по платежам
    for p in payments:
        amount = p.amount or 0
        total_payments += amount
        
        ws_payments.append([
            p.payment_date.strftime('%d.%m.%Y'),
            p.membership.client.full_name,
            p.membership.membership_type.name,
            amount,
            request.user.get_full_name()
        ])
    
    # Добавляем итоги для платежей
    ws_payments.append([])  # Пустая строка
    ws_payments.append([
        'ИТОГО ПЛАТЕЖИ:', '', '',
        total_payments,
        ''
    ])
    
    # --- Лист с общими итогами ---
    ws_summary = wb.create_sheet(title="Общие итоги")
    
    # Заголовок
    ws_summary.append(["ОБЩИЕ ИТОГИ ЗА ПЕРИОД"])
    ws_summary.append([f"С {date_from} по {date_to}"])
    ws_summary.append([])  # Пустая строка
    
    # Данные итогов
    ws_summary.append(["ПОКАЗАТЕЛЬ", "СУММА"])
    ws_summary.append(["Общая стоимость абонементов", total_full_price])
    ws_summary.append(["Общая скидка", total_discount])
    ws_summary.append(["Оплачено при продажах", total_paid_sales])
    ws_summary.append(["Получено платежей", total_payments])
    ws_summary.append(["Общая сумма оплат", total_paid_sales + total_payments])
    ws_summary.append(["Общий долг", total_debt])
    
    # Применяем стили к итогам
    from openpyxl.styles import Font, Alignment
    bold_font = Font(bold=True)
    total_font = Font(bold=True, color="FF0000")
    center_aligned = Alignment(horizontal='center')
    
    # Заголовок
    ws_summary.merge_cells('A1:B1')
    title_cell = ws_summary['A1']
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = center_aligned
    
    # Период
    ws_summary.merge_cells('A2:B2')
    period_cell = ws_summary['A2']
    period_cell.alignment = center_aligned
    
    # Итоговые суммы
    for row in range(4, 10):
        ws_summary[f'A{row}'].font = bold_font
        ws_summary[f'B{row}'].font = bold_font
    
    # Общая сумма оплат
    ws_summary['B8'].font = total_font
    
    # Настраиваем ширину столбцов
    for col in ['A', 'B']:
        ws_summary.column_dimensions[col].width = 30
    
    # Формируем ответ
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=sales_report.xlsx'
    wb.save(response)
    return response
# def export_sales_report(request):
#     # Получаем параметры фильтрации
#     date_from = request.GET.get('date_from')
#     date_to = request.GET.get('date_to')
    
#     # Фильтруем данные
#     memberships = Membership.objects.filter(
#         purchase_date__gte=date_from,
#         purchase_date__lte=date_to
#     ).select_related('client', 'membership_type')
    
#     # Создаем Excel файл
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Отчет по продажам"
    
#     # Заголовки
#     headers = [
#         'Дата', 'Клиент', 'Тип абонемента', 
#         'Полная стоимость', 'Скидка', 'Оплачено', 
#         'Долг', 'Ответственный'
#     ]
#     ws.append(headers)
    
#     # Инициализируем переменные для сумм
#     total_full_price = 0
#     total_discount = 0
#     total_paid = 0
#     total_debt = 0
    
#     # Собираем все данные сначала
#     all_data = []
#     for m in memberships:
#         # Получаем числовые значения
#         full_price = m.membership_type.price or 0
#         discount = m.discount_amount or 0
#         paid = m.paid_amount or 0
#         debt = m.debt_amount or 0
        
#         # Суммируем значения
#         total_full_price += full_price
#         total_discount += discount
#         total_paid += paid
#         total_debt += debt
        
#         all_data.append([
#             m.purchase_date.strftime('%d.%m.%Y'),
#             m.client.full_name,
#             m.membership_type.name,
#             full_price,
#             discount,
#             paid,
#             debt,
#             request.user.get_full_name()
#         ])
    
#     # Записываем данные с разбивкой на страницы
#     ROWS_PER_PAGE = 20  # Количество строк на одной странице
#     current_row_count = 0
    
#     for row in all_data:
#         ws.append(row)
#         current_row_count += 1
        
#         # Если достигли лимита строк на странице
#         if current_row_count >= ROWS_PER_PAGE:
#             # Добавляем пустую строку для разделения
#             ws.append([])
#             # Сбрасываем счетчик
#             current_row_count = 0
#             # Добавляем заголовки для новой страницы
#             ws.append(headers)
    
#     # Добавляем общий итог
#     ws.append([])  # Пустая строка для разделения
#     ws.append([
#         'ОБЩИЙ ИТОГО:', '', '',
#         total_full_price,
#         total_discount,
#         total_paid,
#         total_debt,
#         ''
#     ])
    
#     # Применяем стиль к итоговой строке
#     from openpyxl.styles import Font
#     bold_font = Font(bold=True)
    
#     last_row = ws.max_row
#     for col in range(1, 9):  # Для всех столбцов от A до H
#         cell = ws.cell(row=last_row, column=col)
#         cell.font = bold_font
    
#     # Настраиваем печать
#     ws.print_options.horizontalCentered = True
#     ws.print_options.verticalCentered = True
#     ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
#     ws.page_setup.fitToHeight = False
#     ws.page_setup.fitToWidth = True
    
#     # Формируем ответ
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = 'attachment; filename=sales_report.xlsx'
#     wb.save(response)
#     return response
# def export_sales_report(request):
#     # Получаем параметры фильтрации
#     date_from = request.GET.get('date_from')
#     date_to = request.GET.get('date_to')
    
#     # Фильтруем данные
#     memberships = Membership.objects.filter(
#         purchase_date__gte=date_from,
#         purchase_date__lte=date_to
#     ).select_related('client', 'membership_type')
    
#     # Создаем Excel файл
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Отчет по продажам"
    
#     # Заголовки
#     headers = [
#         'Дата', 'Клиент', 'Тип абонемента', 
#         'Полная стоимость', 'Скидка', 'Оплачено', 
#         'Долг', 'Ответственный'
#     ]
#     ws.append(headers)
    
#     # Настройки для разбивки на страницы
#     ROWS_PER_PAGE = 20  # Количество строк на одной странице
#     current_row_count = 0
#     page_counter = 1
    
#     # Переменные для сумм
#     total_full_price = 0
#     total_discount = 0
#     total_paid = 0
#     total_debt = 0
    
#     # Переменные для промежуточных сумм
#     page_full_price = 0
#     page_discount = 0
#     page_paid = 0
#     page_debt = 0

#     for m in memberships:
#         # Получаем числовые значения
#         full_price = m.membership_type.price or 0
#         discount = m.discount_amount or 0
#         paid = m.paid_amount or 0
#         debt = m.debt_amount or 0
        
#         # Обновляем суммы
#         total_full_price += full_price
#         total_discount += discount
#         total_paid += paid
#         total_debt += debt
        
#         page_full_price += full_price
#         page_discount += discount
#         page_paid += paid
#         page_debt += debt
        
#         # Добавляем строку с данными
#         ws.append([
#             m.purchase_date.strftime('%d.%m.%Y'),
#             m.client.full_name,
#             m.membership_type.name,
#             full_price,
#             discount,
#             paid,
#             debt,
#             request.user.get_full_name()
#         ])
        
#         current_row_count += 1
        
#         # Добавляем промежуточный итог для страницы
#         if current_row_count >= ROWS_PER_PAGE:
#             # Итог для текущей страницы
#             ws.append([
#                 f'ИТОГО Страница {page_counter}:', '', '',
#                 page_full_price,
#                 page_discount,
#                 page_paid,
#                 page_debt,
#                 ''
#             ])
            
#             # Сброс счетчиков для новой страницы
#             current_row_count = 0
#             page_counter += 1
#             page_full_price = 0
#             page_discount = 0
#             page_paid = 0
#             page_debt = 0
            
#             # Добавляем разделитель
#             ws.append([])
#             ws.append(headers)  # Повторяем заголовки для новой страницы

#     # Добавляем промежуточный итог для последней страницы
#     if current_row_count > 0:
#         ws.append([
#             f'ИТОГО Страница {page_counter}:', '', '',
#             page_full_price,
#             page_discount,
#             page_paid,
#             page_debt,
#             ''
#         ])
    
#     # Добавляем общий итог
#     ws.append([])
#     ws.append([
#         'ОБЩИЙ ИТОГО:', '', '',
#         total_full_price,
#         total_discount,
#         total_paid,
#         total_debt,
#         ''
#     ])
    
#     # Настраиваем печать для многостраничных отчетов
#     ws.print_options.horizontalCentered = True
#     ws.print_options.verticalCentered = True
#     ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
#     ws.sheet_properties.pageSetUpPr.fitToPage = True
    
#     # Формируем ответ
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = 'attachment; filename=sales_report.xlsx'
#     wb.save(response)
#     return response



# def export_sales_report(request):
#     # Получаем параметры фильтрации
#     date_from = request.GET.get('date_from')
#     date_to = request.GET.get('date_to')
    
#     # Фильтруем данные
#     memberships = Membership.objects.filter(
#         purchase_date__gte=date_from,
#         purchase_date__lte=date_to
#     ).select_related('client', 'membership_type')
    
#     # Создаем Excel файл
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Отчет по продажам"
    
#     # Заголовки
#     headers = [
#         'Дата', 'Клиент', 'Тип абонемента', 
#         'Полная стоимость', 'Скидка', 'Оплачено', 
#         'Долг', 'Ответственный'
#     ]
#     ws.append(headers)
    
#     # Инициализируем переменные для сумм
#     total_full_price = 0
#     total_discount = 0
#     total_paid = 0
#     total_debt = 0
    
#     # Данные
#     for m in memberships:
#         # Получаем числовые значения (защита от None)
#         full_price = m.membership_type.price or 0
#         discount = m.discount_amount or 0
#         paid = m.paid_amount or 0
#         debt = m.debt_amount or 0
        
#         # Суммируем значения
#         total_full_price += full_price
#         total_discount += discount
#         total_paid += paid
#         total_debt += debt
        
#         ws.append([
#             m.purchase_date.strftime('%d.%m.%Y'),
#             m.client.full_name,
#             m.membership_type.name,
#             full_price,
#             discount,
#             paid,
#             debt,
#             request.user.get_full_name()
#         ])
    
#     # Добавляем строку с итогами
#     ws.append([
#         'ИТОГО:', '', '',
#         total_full_price,
#         total_discount,
#         total_paid,
#         total_debt,
#         ''
#     ])
    
#     # Формируем ответ
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = 'attachment; filename=sales_report.xlsx'
#     wb.save(response)
#     return response


# def export_sales_report(request):
#     # Получаем параметры фильтрации
#     date_from = request.GET.get('date_from')
#     date_to = request.GET.get('date_to')
    
#     # Фильтруем данные
#     memberships = Membership.objects.filter(
#         purchase_date__gte=date_from,
#         purchase_date__lte=date_to
#     ).select_related('client', 'membership_type')
    
#     # Создаем Excel файл
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Отчет по продажам"
    
#     # Заголовки
#     headers = [
#         'Дата', 'Клиент', 'Тип абонемента', 
#         'Полная стоимость', 'Скидка', 'Оплачено', 
#         'Долг', 'Ответственный'
#     ]
#     ws.append(headers)
    
#     # Данные
#     for m in memberships:
#         ws.append([
#             m.purchase_date.strftime('%d.%m.%Y'),
#             m.client.full_name,
#             m.membership_type.name,
#             m.membership_type.price,
#             m.discount_amount,
#             m.paid_amount,
#             m.debt_amount,
#             request.user.get_full_name()  # или другой способ получить ответственного
#         ])
    
#     # Формируем ответ
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = 'attachment; filename=sales_report.xlsx'
#     wb.save(response)
#     return response
# def export_reports(request):
#     period = request.GET.get('period', 'day')
#     date_from = request.GET.get('date_from')
#     date_to = request.GET.get('date_to')
    
#     # Фильтрация данных по периоду
#     # ... логика фильтрации ...
    
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Отчет"
    
#     # Заполнение данными
#     # ... логика заполнения Excel ...
    
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = 'attachment; filename=report.xlsx'
#     wb.save(response)
#     return response
from datetime import datetime
@login_required
@transaction.atomic
def withdraw_cash(request):
    cash_register = get_main_cash_register()
    
    if request.method == 'POST':
        form = WithdrawForm(request.POST)
        if form.is_valid():
            amount = form.cleaned_data['amount']
            reason = form.cleaned_data['reason']
            
            if amount > cash_register.balance:
                messages.error(request, 'Недостаточно средств в кассе!')
                return redirect('sell_card')
             # Создаем запись о платеже
            payment1 = Payment.objects.create(
              
                amount=-amount,
                payment_date = timezone.now(), #datetime.day()
                is_debt_payment=True,
                notes=f"Изьятие денег из кассы. Причина: {reason}",
                created_by=request.user
            )
            # Создаем операцию изъятия
            CashOperation.objects.create(
                payment= payment1,
                cash_register=cash_register,
                operation_type='outcome',
                amount=amount,
                notes=f"Изъятие из кассы. Причина: {reason}",
                created_by=request.user
            )
            
            messages.success(request, f'Из кассы изъято {amount} сом')
            return redirect('sell_card')
    
    return redirect('sell_card')


from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from .forms import OneTimeMembershipForm
from .models import Client, Payment, CashRegister, CashOperation
from django.shortcuts import get_object_or_404, render, redirect
from .models import Client
from .forms import ClientEditForm

def edit_client(request, client_id):
    client = get_object_or_404(Client, pk=client_id)
    
    if request.method == 'POST':
        form = ClientEditForm(request.POST, request.FILES, instance=client)
        if form.is_valid():
            form.save()
            messages.success(request, 'Данные клиента успешно обновлены!')
            return redirect('client_detail',client_id)  # Перенаправление после успеха
    else:
        form = ClientEditForm(instance=client)
    
    return render(request, 'core/edit_client.html', {
        'form': form,
        'client': client
    })
@login_required
def sell_one_time_membership(request):
    if request.method == 'POST':
        form = OneTimeMembershipForm(request.POST)
        if form.is_valid():
            # Рассчитываем цену в зависимости от времени
            price = form.cleaned_data['price']
#             client = Client.objects.filter(phone=form.cleaned_data['phone']).first()
# if client:
#     # Обновляем данные существующего клиента
# else:
#     # Создаем нового
            # Создаем или находим клиента
            client, created = Client.objects.get_or_create(
                phone=form.cleaned_data['phone'],
                defaults={
                    'full_name': form.cleaned_data['client_full_name'],
                    'created_by': request.user
                },
                notes = f"Купил одноразовый абонимент в {datetime.now()}"
            )
            
            if not created:
                client.full_name = form.cleaned_data['client_full_name']
                client.save()
            
            # Создаем платеж
            payment = Payment.objects.create(
                amount=price,
                payment_method=form.cleaned_data['payment_method'],
                notes=f"Одноразовый абонемент ({'до 14:00' if price == 200 else 'после 14:00'})имя:{form.cleaned_data['client_full_name']} {form.cleaned_data['notes']}",
                created_by=request.user
            )
            
            # Записываем операцию в кассу
            cash_register = CashRegister.objects.first()  # Или ваша логика выбора кассы
            if cash_register:
                CashOperation.objects.create(
                    cash_register=cash_register,
                    operation_type='income',
                    amount=price,
                    payment=payment,
                    notes=f"Оплата одноразового абонемента",
                    created_by=request.user
                )
            
            messages.success(request, f'Одноразовый абонемент продан за {price} сом')
            return redirect('client_list')
    else:
        form = OneTimeMembershipForm()
    
    return render(request, 'sell_one_time_membership.html', {
        'form': form,
        # 'current_price': form.calculate_price()
    })
    
